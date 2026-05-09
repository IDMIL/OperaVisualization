import html as html_module
import io
import zipfile
import openpyxl
import warnings
from os import listdir
from difflib import SequenceMatcher

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    _RICH_TEXT_AVAILABLE = True
except ImportError:
    _RICH_TEXT_AVAILABLE = False

warnings.filterwarnings('ignore')

# openpyxl only supports transitional OOXML namespaces; files saved with
# strict conformance use purl.oclc.org URIs instead. Translate them in-memory.
_STRICT_TO_TRANSITIONAL = {
    b'http://purl.oclc.org/ooxml/spreadsheetml/main':
        b'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    b'http://purl.oclc.org/ooxml/officeDocument/relationships':
        b'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    b'http://purl.oclc.org/ooxml/drawingml/2006/main':
        b'http://schemas.openxmlformats.org/drawingml/2006/main',
}

def load_workbook_strict(path):
    with open(path, 'rb') as f:
        raw = f.read()
    with zipfile.ZipFile(io.BytesIO(raw)) as zin:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith('.xml') or item.filename.endswith('.rels'):
                    for old, new in _STRICT_TO_TRANSITIONAL.items():
                        data = data.replace(old, new)
                zout.writestr(item, data)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, rich_text=True)

def isTypoOf(a, b):
    similarity_ratio = SequenceMatcher(None, a, b).ratio()
    return similarity_ratio >= 0.6

def parseAnnotationCodes(codecell):
    annotation_codes = ('dy', 'du', 'for', 'int', 'mo', 'tim', 'graph')
    codes = [i.strip('-') for i in str(codecell).split()]
    good_codes = []
    for c in codes:
        if c in annotation_codes:
            good_codes.append(c)
        elif '-' in c:
            for subcode in c.split('-'):
                if subcode in annotation_codes:
                    good_codes.append(subcode)
        else:
            for annotation_code in annotation_codes:
                if isTypoOf(annotation_code, c.lower()):
                    good_codes.append(annotation_code)
    return good_codes

def replaceSymbols(line):
    flat = "♭"
    sharp = "♯"
    note_names = ('Do', 'Ré', 'Mi', 'Fa', 'Sol', 'La', 'Si')

    parts = line.split(' ')
    for i, p in enumerate(parts):
        prev_is_notename = i > 1 and parts[i-1].strip("(") in note_names
        if p.startswith("#") and prev_is_notename:
            parts[i] = sharp + p[1:]
        elif p.startswith("b") and prev_is_notename and (len(p) == 1 or not p[1].isalpha()):
            parts[i] = flat + p[1:]
    return ' '.join(parts)

def getStartMeasures(a):
    with open(f'act{a}pagebars.txt', 'r') as f:
        lines = f.readlines()
    return [int(l.split("#")[0].strip().split(".")[0]) for l in lines]

start_measures = [getStartMeasures(1) + [717], getStartMeasures(2) + [818], getStartMeasures(3) + [392]]
firstpages = [5, 174, 381]

def pageToBarRange(page: int):
    for act in range(len(firstpages)-1, -1, -1):
        if page >= firstpages[act]:
            return (start_measures[act][page - firstpages[act]],
                    start_measures[act][page + 1 - firstpages[act]] - 1)

def cell_to_str(value):
    if value is None:
        return ''
    if isinstance(value, float):
        return str(int(value))
    return str(value)

def cell_to_html(cell):
    """Return cell text as an HTML string, wrapping italic runs in <i> tags."""
    value = cell.value
    if value is None:
        return ''

    if _RICH_TEXT_AVAILABLE and isinstance(value, CellRichText):
        parts = []
        for block in value:
            if isinstance(block, TextBlock):
                text = html_module.escape(replaceSymbols(str(block.text)), quote=False)
                if block.font and block.font.i:
                    parts.append(f'<i>{text}</i>')
                else:
                    parts.append(text)
            else:
                parts.append(html_module.escape(replaceSymbols(str(block)), quote=False))
        return ''.join(parts)

    # Uniform-formatted cell: check cell-level italic
    if isinstance(value, float):
        text = str(int(value))
    else:
        text = str(value)
    text = html_module.escape(replaceSymbols(text), quote=False)
    if cell.font and cell.font.i:
        return f'<i>{text}</i>'
    return text

def parse_annotations_sheet(ws, act_number, isGeneral):
    annotations = []
    current_measures = None
    current_page_range = None
    is_header = True
    for row in ws.iter_rows():
        if is_header:
            is_header = False
            continue

        if row[2].value is None or cell_to_str(row[2].value).strip() == "":
            continue

        a = {}
        a['code'] = parseAnnotationCodes(cell_to_str(row[0].value))

        cell1 = cell_to_str(row[1].value)

        if isGeneral and cell1 != '':
            page_range_raw = cell1.split('-')
            try:
                page_range_parsed = [int(i) for i in page_range_raw]
            except ValueError:
                print("cannot parse page range", page_range_raw, "on row", row)
                page_range_parsed = current_page_range
            if len(page_range_parsed) == 1:
                measure_range = pageToBarRange(page_range_parsed[0])
                current_measures = [measure_range[0], measure_range[1]]
                current_page_range = [page_range_parsed[0], page_range_parsed[0]]
            else:
                current_measures = [pageToBarRange(page_range_parsed[0])[0], pageToBarRange(page_range_parsed[1])[1]]
                current_page_range = page_range_parsed

        elif cell1 != '':
            measure_range_raw = cell1.split('-')
            try:
                measure_range = [int(i) for i in measure_range_raw]
            except ValueError:
                print("cannot parse measure range", measure_range_raw, "on row", row)
                measure_range = current_measures
            if len(measure_range) == 1:
                current_measures = [measure_range[0], measure_range[0]]
            else:
                if len(str(measure_range[1])) < len(str(measure_range[0])):
                    # Number format like 123-35
                    prefix = str(measure_range[0])[:-len(str(measure_range[1]))]
                    measure_range[1] = int(prefix + str(measure_range[1]))
                current_measures = [measure_range[0], measure_range[1]]

        a['annotation'] = cell_to_html(row[2])
        a['act'] = act_number
        a['is_general'] = isGeneral
        a['page_range'] = current_page_range if isGeneral else [0, 0]
        a['measure_range'] = current_measures
        annotations.append(a)

    return annotations

all_annotations = []
for f in listdir("../annotations"):
    if not f.endswith('.xlsx'):
        continue

    act_number = 0
    if 'act1' in f.lower():
        act_number = 1
    elif 'actiii' in f.lower():
        act_number = 3
    else:
        print("Cannot determine act of", f)
        continue

    wb = load_workbook_strict('../annotations/' + f)

    if not wb.sheetnames:
        print("No sheets found in", f, "- skipping")
        continue

    for sheet_name in wb.sheetnames:
        isGeneral = 'mesur' not in sheet_name.lower()
        ws = wb[sheet_name]
        all_annotations += parse_annotations_sheet(ws, act_number, isGeneral)

all_annotations.sort(key=lambda a: a['act'] * 10000 + a['measure_range'][0])
with open("../site/src/data/annotations.ts", 'w', encoding='utf8') as annotations_file:
    annotations_file.write("""export type AnnotationCode = 'dy' | 'du' | 'for' | 'int' | 'mo' | 'tim' | 'graph';

export interface Annotation {
    code : Array<AnnotationCode>;
    annotation : string;
    act : number;
    is_general: boolean;
    page_range: [number, number];
    measure_range : [number, number];
}

export const annotations : Array<Annotation> =
""")
    annotations_file.write(str(all_annotations)
                           .replace("'is_general': False", "'is_general': false")
                           .replace("'is_general': True", "'is_general': true"))
